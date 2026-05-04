"""Excel ingestion: participantes_lista from professor's spreadsheet, CIIU catalog from DANE."""
import hashlib
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from openpyxl import load_workbook

from ..auth import require_internal_token
from ..db import get_supabase_admin

router = APIRouter(prefix="/ingest", tags=["ingest"], dependencies=[Depends(require_internal_token)])


def _sha256(s: str) -> str:
    return hashlib.sha256(s.strip().encode()).hexdigest()


@router.post("/participantes")
async def ingest_participantes(
    cohorte_id: str = Form(...),
    file: UploadFile = File(...),
):
    """Excel columns: nombre_completo, cedula, email."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "File must be xlsx/xls")

    content = await file.read()
    wb = load_workbook(BytesIO(content), read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip().lower() if c else "" for c in next(rows, [])]

    required = {"nombre_completo", "cedula", "email"}
    missing = required - set(header)
    if missing:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Missing columns: {missing}")

    idx = {h: header.index(h) for h in required}

    sb = get_supabase_admin()
    # Verify cohorte exists
    coh = sb.table("cohortes").select("id").eq("id", cohorte_id).maybeSingle().execute()
    if not coh.data:
        raise HTTPException(status.HTTP_404_NOT_FOUND, f"Cohorte {cohorte_id} not found")

    inserted, errors = 0, []
    for i, row in enumerate(rows, start=2):
        try:
            nombre = str(row[idx["nombre_completo"]]).strip()
            cedula = str(row[idx["cedula"]]).strip().replace(".", "").replace("-", "").replace(" ", "")
            email = str(row[idx["email"]]).strip().lower()
            if not (nombre and cedula and email):
                continue
            # Backend Node will encrypt later. For now we store hash + plaintext-encrypted via SQL function
            # Or skip encryption here and let backend do it. For MVP we mark as TODO.
            sb.table("participantes_lista").upsert({
                "cohorte_id": cohorte_id,
                "nombre_completo": nombre,
                "cedula_encriptada": cedula,  # TODO: encrypt
                "cedula_hash": _sha256(cedula),
                "email_encriptado": email,    # TODO: encrypt
                "email_hash": _sha256(email),
                "estado": "pendiente_activacion",
            }, on_conflict="cohorte_id,cedula_hash").execute()
            inserted += 1
        except Exception as e:
            errors.append({"row": i, "error": str(e)})

    return {"inserted": inserted, "errors": errors[:20]}
